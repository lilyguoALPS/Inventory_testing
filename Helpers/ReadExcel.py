import openpyxl


def ReadData(file_path):
    data_list = []
    path = file_path
    workbook = openpyxl.load_workbook(path,data_only=True)
    sheet  = workbook["Sheet1"]   #workbook.get_sheet_by_name("Sheet1")
    rows = sheet.max_row
    cols = sheet.max_column
    
    for r in range(2,rows + 1):
        tuple_data = (sheet.cell(r,1).value,)
        for c in range(2,cols + 1):
            var_r_c = sheet.cell(r,c).value
            tuple_data += (var_r_c,)
        data_list.append(tuple_data)
    return data_list
